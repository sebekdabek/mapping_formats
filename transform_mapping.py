import csv
import json
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Source Entity", "Source Column",
    "EDM Entity", "EDM Column",
    "Target Entity", "Target Column"
]

DATA = [
    ["CRM_Customers",    "cust_id",          "Party",          "party_id",         "DWH_Dim_Customer",    "customer_key"],
    ["CRM_Customers",    "first_name",        "Party",          "given_name",       "DWH_Dim_Customer",    "first_name"],
    ["CRM_Customers",    "last_name",         "Party",          "family_name",      "DWH_Dim_Customer",    "last_name"],
    ["CRM_Customers",    "email_address",     "Party",          "contact_email",    "DWH_Dim_Customer",    "email"],
    ["CRM_Customers",    "phone_number",      "Party",          "contact_phone",    "DWH_Dim_Customer",    "phone"],
    ["CRM_Customers",    "country_code",      "Party",          "country_iso2",     "DWH_Dim_Customer",    "country_code"],
    ["ERP_Orders",       "order_id",          "SalesOrder",     "order_id",         "DWH_Fact_Sales",      "order_key"],
    ["ERP_Orders",       "order_date",        "SalesOrder",     "transaction_date", "DWH_Fact_Sales",      "order_date"],
    ["ERP_Orders",       "cust_ref",          "SalesOrder",     "party_ref",        "DWH_Fact_Sales",      "customer_key"],
    ["ERP_Orders",       "total_amount",      "SalesOrder",     "gross_amount",     "DWH_Fact_Sales",      "gross_sales"],
    ["ERP_Orders",       "discount_amt",      "SalesOrder",     "discount_amount",  "DWH_Fact_Sales",      "discount_amount"],
    ["ERP_Orders",       "net_amount",        "SalesOrder",     "net_amount",       "DWH_Fact_Sales",      "net_sales"],
    ["ERP_Orders",       "currency_code",     "SalesOrder",     "currency_iso3",    "DWH_Fact_Sales",      "currency_code"],
    ["ERP_Orders",       "status_code",       "SalesOrder",     "order_status",     "DWH_Fact_Sales",      "order_status"],
    ["ERP_Products",     "prod_id",           "Product",        "product_id",       "DWH_Dim_Product",     "product_key"],
    ["ERP_Products",     "prod_name",         "Product",        "product_name",     "DWH_Dim_Product",     "product_name"],
    ["ERP_Products",     "sku_code",          "Product",        "sku",              "DWH_Dim_Product",     "sku"],
    ["ERP_Products",     "category_id",       "Product",        "category_ref",     "DWH_Dim_Product",     "category_key"],
    ["ERP_Products",     "unit_price",        "Product",        "list_price",       "DWH_Dim_Product",     "list_price"],
    ["ERP_Products",     "cost_price",        "Product",        "standard_cost",    "DWH_Dim_Product",     "standard_cost"],
    ["HR_Employees",     "emp_id",            "Person",         "person_id",        "DWH_Dim_Employee",    "employee_key"],
    ["HR_Employees",     "full_name",         "Person",         "full_name",        "DWH_Dim_Employee",    "employee_name"],
    ["HR_Employees",     "dept_code",         "Person",         "org_unit_ref",     "DWH_Dim_Employee",    "department_key"],
    ["HR_Employees",     "hire_date",         "Person",         "start_date",       "DWH_Dim_Employee",    "hire_date"],
    ["HR_Employees",     "job_title",         "Person",         "role_title",       "DWH_Dim_Employee",    "job_title"],
    ["FIN_Invoices",     "invoice_id",        "Invoice",        "invoice_id",       "DWH_Fact_Finance",    "invoice_key"],
    ["FIN_Invoices",     "invoice_date",      "Invoice",        "issue_date",       "DWH_Fact_Finance",    "invoice_date"],
    ["FIN_Invoices",     "due_date",          "Invoice",        "payment_due_date", "DWH_Fact_Finance",    "due_date"],
    ["FIN_Invoices",     "invoice_amount",    "Invoice",        "total_amount",     "DWH_Fact_Finance",    "invoice_amount"],
    ["FIN_Invoices",     "paid_flag",         "Invoice",        "is_paid",          "DWH_Fact_Finance",    "is_paid"],
]


def export_csv(path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(DATA)


def export_json(path):
    records = [dict(zip(HEADERS, row)) for row in DATA]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)


def export_yaml(path):
    records = [dict(zip(HEADERS, row)) for row in DATA]
    with open(path, "w", encoding="utf-8") as f:
        yaml.dump(records, f, allow_unicode=True, default_flow_style=False, sort_keys=False)


def export_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Mapping"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="2E5339")  # dark forest green
    cell_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="EEF3F0")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [18, 22, 16, 22, 22, 22]

    # Header row
    for col_idx, (header, width) in enumerate(zip(HEADERS, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, row in enumerate(DATA, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = center if col_idx in (3, 5) else left
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(path)


if __name__ == "__main__":
    import os
    out = os.path.dirname(os.path.abspath(__file__))
    os.makedirs(out, exist_ok=True)

    export_csv(f"{out}/data_mapping.csv")
    export_json(f"{out}/data_mapping.json")
    export_yaml(f"{out}/data_mapping.yaml")
    export_xlsx(f"{out}/data_mapping.xlsx")

    print("Done — 4 files written to outputs/")
